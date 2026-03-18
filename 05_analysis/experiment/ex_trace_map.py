#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据 runsheet + 实验设计 + samples.xlsx 生成 EX × batch 的路径图、明细表与 QA 报表。

核心思路：
1) 从 runsheet 读取 action / day / id
2) 将 id 解析为 sample_id 或 mice_id
3) 使用 samples.xlsx 的 samples sheet 将 sample_id / mice_id 关联到 experiment_id
4) 以 experiment_id + batch_label 为单位，汇总该批样品经历过的 action
5) 输出：
   - timeline.png          左→右路径图
   - action_log.xlsx       action_log / batch_summary / qa 三个 sheet
   - action_log.csv        长表，便于后续脚本再利用

用法示例：
python 05_analysis/experiment/ex_trace_map.py \
  --runsheet 09_records/20260314_runsheet.yaml \
  --design 09_records/20260314_实验设计.yaml \
  --samples 03_metarials/sample/samples.xlsx \
  --outdir 07_results/experiment
"""

from __future__ import annotations

import argparse
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import matplotlib
matplotlib.use("Agg")

import matplotlib.pyplot as plt
from matplotlib import rcParams

# 尽量使用容器中常见的 CJK 字体，避免中文显示为方框
rcParams["font.family"] = "sans-serif"
rcParams["font.sans-serif"] = [
    "Noto Sans CJK SC",
    "Noto Sans CJK JP",
    "Noto Sans CJK KR",
    "Arial Unicode MS",
    "DejaVu Sans",
]
rcParams["axes.unicode_minus"] = False
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =========================
# 基础解析函数
# =========================

SAMPLE_RE = re.compile(r"^([A-Za-z])(\d+)$")
RANGE_RE = re.compile(r"^([A-Za-z])(\d+)\s*-\s*([A-Za-z]?)(\d+)$")


@dataclass
class StepRecord:
    day: int
    action: str
    action_order: int
    id_spec_raw: str
    id_mode: str  # sample | mouse | unknown
    entity_id: str  # sample_id 或 mice_id


def normalize_identifier(token: str) -> str:
    token = str(token).strip()
    token = token.replace("，", ",").replace("；", ";")
    token = re.sub(r"\s+", "", token)
    return token.upper()


def parse_range_token(token: str) -> List[str]:
    """解析单个 token，可以是单个 ID 或区间。支持 S0137-S0152 / M009-M016 / s00153-s0158。"""
    token = normalize_identifier(token)
    if not token:
        return []

    m = SAMPLE_RE.match(token)
    if m:
        prefix, digits = m.groups()
        width = len(digits)
        return [f"{prefix}{int(digits):0{width}d}"]

    m = RANGE_RE.match(token)
    if m:
        p1, d1, p2, d2 = m.groups()
        p2 = p2 or p1
        if p1 != p2:
            raise ValueError(f"不支持跨前缀区间: {token}")
        n1, n2 = int(d1), int(d2)
        if n2 < n1:
            raise ValueError(f"区间终点小于起点: {token}")
        width = max(len(d1), len(d2), 3)
        return [f"{p1}{n:0{width}d}" for n in range(n1, n2 + 1)]

    raise ValueError(f"无法识别的 ID token: {token}")


def split_mixed_id_spec(spec: str) -> List[str]:
    spec = str(spec).strip()
    if not spec:
        return []
    parts = [x.strip() for x in spec.replace("；", ",").replace("，", ",").split(",")]
    return [p for p in parts if p]


def infer_id_mode(expanded_ids: Sequence[str]) -> str:
    if not expanded_ids:
        return "unknown"
    prefixes = {x[0].upper() for x in expanded_ids if x}
    if prefixes == {"S"}:
        return "sample"
    if prefixes == {"M"}:
        return "mouse"
    return "unknown"


def expand_id_spec(id_spec: str | List[str]) -> Tuple[List[str], List[str]]:
    """返回 (expanded_ids, parse_errors)"""
    specs = id_spec if isinstance(id_spec, list) else [id_spec]
    expanded: List[str] = []
    errors: List[str] = []

    for raw in specs:
        for token in split_mixed_id_spec(str(raw)):
            try:
                expanded.extend(parse_range_token(token))
            except Exception as e:  # noqa: BLE001
                errors.append(f"{token}: {e}")
    return sorted(set(expanded)), errors


# =========================
# 文件读取
# =========================

def load_design_map(design_path: Path) -> Tuple[Dict[str, dict], List[str]]:
    with open(design_path, "r", encoding="utf-8") as f:
        design = yaml.safe_load(f) or {}
    experiments = design.get("experiments", [])
    design_map: Dict[str, dict] = {}
    order: List[str] = []
    for item in experiments:
        ex_id = str(item.get("id", "")).strip()
        if not ex_id:
            continue
        design_map[ex_id] = item
        order.append(ex_id)
    return design_map, order


def load_samples_table(samples_path: Path) -> pd.DataFrame:
    df = pd.read_excel(samples_path, sheet_name="samples")
    df.columns = [str(c).strip() if c is not None else "" for c in df.columns]

    rename_map = {}
    for col in df.columns:
        low = col.lower()
        if low == "experiments_id":
            rename_map[col] = "experiment_id"
        elif low == "sample_id":
            rename_map[col] = "sample_id"
        elif low == "mice_id":
            rename_map[col] = "mice_id"
        elif low == "sample_type":
            rename_map[col] = "sample_type"
        elif low == "treatment":
            rename_map[col] = "treatment"
        elif low == "stimulate":
            rename_map[col] = "stimulate"
        elif low == "genotype":
            rename_map[col] = "genotype"

    df = df.rename(columns=rename_map)
    required = ["sample_id", "mice_id", "experiment_id"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"samples sheet 缺少必要列: {missing}")

    df["sample_id"] = df["sample_id"].astype(str).str.strip().str.upper()
    df["mice_id"] = df["mice_id"].astype(str).str.strip().str.upper()
    df["experiment_id"] = df["experiment_id"].astype(str).str.strip()
    df = df[df["sample_id"].ne("NAN") & df["sample_id"].ne("")].copy()
    return df


def load_runsheet_steps(runsheet_path: Path) -> Tuple[List[StepRecord], List[str]]:
    with open(runsheet_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    steps = data.get("steps", [])
    records: List[StepRecord] = []
    parse_errors: List[str] = []

    for idx, step in enumerate(steps, start=1):
        day = int(step.get("day"))
        action = str(step.get("action", "")).strip()
        raw_id = step.get("id", "")
        raw_specs = raw_id if isinstance(raw_id, list) else [raw_id]
        for raw_piece in raw_specs:
            expanded_ids, errors = expand_id_spec(raw_piece)
            if errors:
                parse_errors.extend([f"step#{idx} action={action}: {e}" for e in errors])
            id_mode = infer_id_mode(expanded_ids)
            records.append(
                StepRecord(
                    day=day,
                    action=action,
                    action_order=idx,
                    id_spec_raw=str(raw_piece),
                    id_mode=id_mode,
                    entity_id="|".join(expanded_ids),
                )
            )
    return records, parse_errors


# =========================
# 数据整形
# =========================

def contiguous_blocks(sample_ids: Sequence[str]) -> List[List[str]]:
    if not sample_ids:
        return []
    nums = sorted({int(s[1:]) for s in sample_ids})
    blocks: List[List[int]] = [[nums[0]]]
    for n in nums[1:]:
        if n == blocks[-1][-1] + 1:
            blocks[-1].append(n)
        else:
            blocks.append([n])
    return [[f"S{x:04d}" for x in block] for block in blocks]


def block_label(sample_ids: Sequence[str]) -> str:
    ids = sorted(sample_ids, key=lambda x: int(x[1:]))
    if not ids:
        return "EMPTY"
    if len(ids) == 1:
        return ids[0]
    return f"{ids[0]}-{ids[-1]}"


def build_action_log(
    step_records: List[StepRecord],
    sample_df: pd.DataFrame,
    design_map: Dict[str, dict],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    sample_lookup = sample_df.set_index("sample_id", drop=False)
    by_mouse = sample_df.groupby("mice_id")["sample_id"].apply(list).to_dict()

    parsed_rows: List[dict] = []
    qa_rows: List[dict] = []

    # 先确定此次 run 里明确出现过的 sample_id 和对应 experiment
    explicit_sample_ids = set()
    for rec in step_records:
        entity_ids = [x for x in rec.entity_id.split("|") if x]
        if rec.id_mode == "sample":
            explicit_sample_ids.update(entity_ids)
    explicit_experiments = set(sample_df[sample_df["sample_id"].isin(explicit_sample_ids)]["experiment_id"].dropna())

    for rec in step_records:
        entity_ids = [x for x in rec.entity_id.split("|") if x]
        if not entity_ids:
            qa_rows.append(
                {
                    "level": "warning",
                    "issue_type": "empty_id_resolution",
                    "day": rec.day,
                    "action": rec.action,
                    "id_spec_raw": rec.id_spec_raw,
                    "detail": "该步骤未解析出任何 ID。",
                }
            )
            continue

        if rec.id_mode == "sample":
            resolved_sample_ids = [sid for sid in entity_ids if sid in sample_lookup.index]
            unresolved = [sid for sid in entity_ids if sid not in sample_lookup.index]
            for sid in unresolved:
                qa_rows.append(
                    {
                        "level": "warning",
                        "issue_type": "sample_not_found",
                        "day": rec.day,
                        "action": rec.action,
                        "id_spec_raw": rec.id_spec_raw,
                        "detail": sid,
                    }
                )

        elif rec.id_mode == "mouse":
            # 将 mouse 步骤映射到本次 run 中涉及的 experiment 对应样品
            resolved_sample_ids = []
            for mid in entity_ids:
                matched = by_mouse.get(mid, [])
                if not matched:
                    qa_rows.append(
                        {
                            "level": "warning",
                            "issue_type": "mouse_not_found",
                            "day": rec.day,
                            "action": rec.action,
                            "id_spec_raw": rec.id_spec_raw,
                            "detail": mid,
                        }
                    )
                    continue
                for sid in matched:
                    ex_id = sample_lookup.loc[sid, "experiment_id"]
                    if (not explicit_experiments) or (ex_id in explicit_experiments):
                        resolved_sample_ids.append(sid)
            resolved_sample_ids = sorted(set(resolved_sample_ids), key=lambda x: int(x[1:]))
        else:
            resolved_sample_ids = []
            qa_rows.append(
                {
                    "level": "warning",
                    "issue_type": "unknown_id_mode",
                    "day": rec.day,
                    "action": rec.action,
                    "id_spec_raw": rec.id_spec_raw,
                    "detail": rec.entity_id,
                }
            )

        if not resolved_sample_ids:
            continue

        step_sample_df = sample_df[sample_df["sample_id"].isin(resolved_sample_ids)].copy()
        for ex_id, ex_df in step_sample_df.groupby("experiment_id"):
            ex_sample_ids = sorted(ex_df["sample_id"].tolist(), key=lambda x: int(x[1:]))
            for block in contiguous_blocks(ex_sample_ids):
                label = block_label(block)
                sub = ex_df[ex_df["sample_id"].isin(block)].copy()
                question = design_map.get(ex_id, {}).get("question", "")
                parsed_rows.append(
                    {
                        "experiment_id": ex_id,
                        "experiment_question": question,
                        "batch_label": label,
                        "batch_n": len(block),
                        "day": rec.day,
                        "action_order": rec.action_order,
                        "action": rec.action,
                        "id_mode": rec.id_mode,
                        "id_spec_raw": rec.id_spec_raw,
                        "sample_ids": ", ".join(sorted(block, key=lambda x: int(x[1:]))),
                        "mice_ids": ", ".join(sorted(set(sub["mice_id"].dropna().astype(str)))),
                        "genotypes": ", ".join(sorted(set(sub.get("genotype", pd.Series(dtype=str)).dropna().astype(str)))),
                        "treatments": ", ".join(sorted(set(sub.get("treatment", pd.Series(dtype=str)).dropna().astype(str)))),
                        "stimulates": ", ".join(sorted(set(sub.get("stimulate", pd.Series(dtype=str)).dropna().astype(str)))),
                        "sample_types": ", ".join(sorted(set(sub.get("sample_type", pd.Series(dtype=str)).dropna().astype(str)))),
                    }
                )

    action_log = pd.DataFrame(parsed_rows)
    qa_df = pd.DataFrame(qa_rows)

    if action_log.empty:
        batch_summary = pd.DataFrame(
            columns=[
                "experiment_id", "experiment_question", "batch_label", "batch_n",
                "actions", "days", "sample_ids", "mice_ids", "sample_types"
            ]
        )
        return action_log, batch_summary, qa_df

    action_log = action_log.sort_values(
        ["experiment_id", "batch_label", "day", "action_order", "action"]
    ).reset_index(drop=True)

    # 去重：同一 batch 同一 action 可能因不同 raw id piece 重复进入，这里合并
    agg_cols = ["experiment_id", "experiment_question", "batch_label", "batch_n", "day", "action_order", "action"]
    action_log = (
        action_log.groupby(agg_cols, dropna=False)
        .agg(
            sample_ids=("sample_ids", lambda x: "; ".join(sorted(set(x)))),
            mice_ids=("mice_ids", lambda x: "; ".join(sorted(set(x)))),
            genotypes=("genotypes", lambda x: "; ".join(sorted(set(x)))),
            treatments=("treatments", lambda x: "; ".join(sorted(set(x)))),
            stimulates=("stimulates", lambda x: "; ".join(sorted(set(x)))),
            sample_types=("sample_types", lambda x: "; ".join(sorted(set(x)))),
            id_modes=("id_mode", lambda x: "; ".join(sorted(set(x)))),
            id_spec_raws=("id_spec_raw", lambda x: " | ".join(sorted(set(str(v) for v in x)))),
        )
        .reset_index()
        .sort_values(["experiment_id", "batch_label", "day", "action_order", "action"])
        .reset_index(drop=True)
    )

    batch_summary = (
        action_log.groupby(["experiment_id", "experiment_question", "batch_label", "batch_n"], dropna=False)
        .agg(
            actions=("action", lambda x: " → ".join(x)),
            days=("day", lambda x: ", ".join(str(v) for v in x)),
            sample_ids=("sample_ids", lambda x: " | ".join(sorted(set(x)))),
            mice_ids=("mice_ids", lambda x: " | ".join(sorted(set(x)))),
            sample_types=("sample_types", lambda x: " | ".join(sorted(set(x)))),
        )
        .reset_index()
    )

    # QA: 设计文件缺失的 experiment
    found_ex = sorted(set(action_log["experiment_id"]))
    missing_design = [ex for ex in found_ex if ex not in design_map]
    if missing_design:
        extra = pd.DataFrame(
            [{
                "level": "warning",
                "issue_type": "experiment_missing_in_design",
                "day": None,
                "action": None,
                "id_spec_raw": None,
                "detail": ex,
            } for ex in missing_design]
        )
        qa_df = pd.concat([qa_df, extra], ignore_index=True)

    # QA: 同一 batch 是否出现时间逆序（严格按 day + order 排后一般不会，但保留检查）
    for (ex_id, batch_label), sub in action_log.groupby(["experiment_id", "batch_label"]):
        seq = list(zip(sub["day"], sub["action_order"]))
        if seq != sorted(seq):
            qa_df = pd.concat(
                [qa_df, pd.DataFrame([{
                    "level": "warning",
                    "issue_type": "non_monotonic_action_order",
                    "day": None,
                    "action": None,
                    "id_spec_raw": None,
                    "detail": f"{ex_id} / {batch_label}",
                }])],
                ignore_index=True,
            )

    if qa_df.empty:
        qa_df = pd.DataFrame([{
            "level": "info",
            "issue_type": "no_issue",
            "day": None,
            "action": None,
            "id_spec_raw": None,
            "detail": "未发现解析级别问题。",
        }])

    return action_log, batch_summary, qa_df


# =========================
# 绘图
# =========================

def build_row_order(batch_summary: pd.DataFrame, design_order: List[str]) -> List[Tuple[str, str]]:
    items = [(r.experiment_id, r.batch_label) for r in batch_summary.itertuples(index=False)]
    by_ex: Dict[str, List[str]] = {}
    for ex_id, batch_label in items:
        by_ex.setdefault(ex_id, []).append(batch_label)

    def batch_sort_key(label: str) -> Tuple[int, int, str]:
        nums = [int(x[1:]) for x in re.findall(r"S\d+", label.upper())]
        if nums:
            return (nums[0], nums[-1], label)
        m = re.findall(r"\d+", label)
        if m:
            return (int(m[0]), int(m[-1]), label)
        return (10**9, 10**9, label)

    design_rank = {ex_id: idx for idx, ex_id in enumerate(design_order)}
    summary_lookup = {
        (r.experiment_id, r.batch_label): r
        for r in batch_summary.itertuples(index=False)
    }

    def endpoint_sort_key(item: Tuple[str, str]) -> Tuple[str, int, int, int, str, str]:
        ex_id, batch_label = item
        row = summary_lookup[(ex_id, batch_label)]
        actions = str(getattr(row, "actions", "") or "")
        action_parts = [x.strip() for x in actions.split("→") if x.strip()]
        endpoint = action_parts[-1] if action_parts else ""
        return (
            endpoint,
            design_rank.get(ex_id, 10**9),
            batch_sort_key(batch_label)[0],
            batch_sort_key(batch_label)[1],
            ex_id,
            batch_label,
        )

    ordered = sorted(set(items), key=endpoint_sort_key)
    return ordered


def draw_timeline(
    action_log: pd.DataFrame,
    batch_summary: pd.DataFrame,
    design_order: List[str],
    output_png: Path,
    title: str,
) -> None:
    if action_log.empty:
        fig, ax = plt.subplots(figsize=(10, 3))
        ax.text(0.5, 0.5, "No action records resolved", ha="center", va="center", fontsize=14)
        ax.axis("off")
        fig.savefig(output_png, dpi=200, bbox_inches="tight")
        plt.close(fig)
        return

    action_axis = (
        action_log[["day", "action_order", "action"]]
        .drop_duplicates()
        .sort_values(["day", "action_order", "action"])
        .reset_index(drop=True)
    )
    action_axis["x"] = range(len(action_axis))
    x_map = {(r.day, r.action_order, r.action): r.x for r in action_axis.itertuples(index=False)}

    row_items = build_row_order(batch_summary, design_order)
    y_map = {item: idx for idx, item in enumerate(row_items)}

    # 尺寸估计：列多就拉宽，行多就拉高
    width = max(14, 1.35 * len(action_axis) + 6)
    height = max(6, 0.55 * len(row_items) + 2.8)
    fig, ax = plt.subplots(figsize=(width, height))

    # 网格背景（不指定具体颜色，只用浅灰默认风格）
    for x in action_axis["x"]:
        ax.axvline(x, linewidth=0.5, alpha=0.2, zorder=0)

    # 按终点步骤分组画轻微分隔
    endpoint_by_row: Dict[Tuple[str, str], str] = {}
    summary_lookup = {
        (r.experiment_id, r.batch_label): r
        for r in batch_summary.itertuples(index=False)
    }
    for item in row_items:
        row = summary_lookup.get(item)
        actions = str(getattr(row, "actions", "") or "") if row is not None else ""
        action_parts = [x.strip() for x in actions.split("→") if x.strip()]
        endpoint_by_row[item] = action_parts[-1] if action_parts else ""

    endpoint_boundaries = []
    prev_endpoint: Optional[str] = None
    for idx, item in enumerate(row_items):
        endpoint = endpoint_by_row.get(item, "")
        if idx > 0 and endpoint != prev_endpoint:
            endpoint_boundaries.append(idx - 0.5)
        prev_endpoint = endpoint
    for y in endpoint_boundaries:
        ax.axhline(y, linewidth=0.8, alpha=0.25, zorder=0)

    # 每一行一条 path
    for (ex_id, batch_label), sub in action_log.groupby(["experiment_id", "batch_label"]):
        if (ex_id, batch_label) not in y_map:
            continue
        y = y_map[(ex_id, batch_label)]
        sub = sub.sort_values(["day", "action_order", "action"])
        xs = [x_map[(r.day, r.action_order, r.action)] for r in sub.itertuples(index=False)]
        ys = [y] * len(xs)
        ax.plot(xs, ys, marker="o", linewidth=2)

    # Y 标签
    y_ticks = list(range(len(row_items)))
    y_labels = []
    for ex_id, batch_label in row_items:
        row = summary_lookup.get((ex_id, batch_label))
        actions = str(getattr(row, "actions", "") or "") if row is not None else ""
        action_parts = [x.strip() for x in actions.split("→") if x.strip()]
        endpoint = action_parts[-1] if action_parts else "NA"
        y_labels.append(f"[{endpoint}] {ex_id} | {batch_label}")
    ax.set_yticks(y_ticks)
    ax.set_yticklabels(y_labels, fontsize=9)

    # X 标签：Day + action
    ax.set_xticks(action_axis["x"])
    x_labels = [f"D{r.day}\n{r.action}" for r in action_axis.itertuples(index=False)]
    ax.set_xticklabels(x_labels, rotation=45, ha="right", fontsize=9)

    ax.set_xlabel("Action timeline")
    ax.set_ylabel("Endpoint-grouped experiment × sample batch")
    ax.set_title(title, fontsize=14)
    ax.invert_yaxis()
    ax.spines[["top", "right"]].set_visible(False)
    plt.tight_layout()
    fig.savefig(output_png, dpi=220, bbox_inches="tight")
    plt.close(fig)


# =========================
# Excel 导出
# =========================

def autosize_worksheet(ws) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            try:
                val = "" if cell.value is None else str(cell.value)
            except Exception:  # noqa: BLE001
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def write_df_to_sheet(ws, df: pd.DataFrame, header_fill: PatternFill) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)


def export_excel(action_log: pd.DataFrame, batch_summary: pd.DataFrame, qa_df: pd.DataFrame, output_xlsx: Path) -> None:
    wb = Workbook()
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    ws1 = wb.active
    ws1.title = "action_log"
    write_df_to_sheet(ws1, action_log, header_fill)

    ws2 = wb.create_sheet("batch_summary")
    write_df_to_sheet(ws2, batch_summary, header_fill)

    ws3 = wb.create_sheet("qa")
    write_df_to_sheet(ws3, qa_df, header_fill)

    wb.save(output_xlsx)


# =========================
# 主流程
# =========================

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate EX × batch action trace map from runsheet/design/samples.xlsx")
    parser.add_argument("--runsheet", required=True, help="runsheet yaml path")
    parser.add_argument("--design", required=True, help="experiment design yaml path")
    parser.add_argument("--samples", required=True, help="samples.xlsx path")
    parser.add_argument("--outdir", required=True, help="output directory")
    parser.add_argument("--title", default=None, help="custom plot title")
    args = parser.parse_args()

    runsheet_path = Path(args.runsheet)
    design_path = Path(args.design)
    samples_path = Path(args.samples)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    design_map, design_order = load_design_map(design_path)
    sample_df = load_samples_table(samples_path)
    step_records, parse_errors = load_runsheet_steps(runsheet_path)

    action_log, batch_summary, qa_df = build_action_log(step_records, sample_df, design_map)

    if parse_errors:
        qa_parse = pd.DataFrame([
            {
                "level": "warning",
                "issue_type": "id_parse_error",
                "day": None,
                "action": None,
                "id_spec_raw": None,
                "detail": msg,
            }
            for msg in parse_errors
        ])
        qa_df = pd.concat([qa_df, qa_parse], ignore_index=True)

    # 补充 experiment_name/label 可读字段
    if not batch_summary.empty:
        batch_summary.insert(
            1,
            "experiment_name_short",
            batch_summary["experiment_question"].apply(lambda x: str(x)[:80] + ("..." if len(str(x)) > 80 else "")),
        )
    if not action_log.empty:
        action_log.insert(
            1,
            "experiment_name_short",
            action_log["experiment_question"].apply(lambda x: str(x)[:80] + ("..." if len(str(x)) > 80 else "")),
        )

    title = args.title or f"EX trace map | {runsheet_path.stem}"
    png_path = outdir / f"{runsheet_path.stem}_timeline.png"
    xlsx_path = outdir / f"{runsheet_path.stem}_action_log.xlsx"
    csv_path = outdir / f"{runsheet_path.stem}_action_log.csv"

    draw_timeline(action_log, batch_summary, design_order, png_path, title)
    export_excel(action_log, batch_summary, qa_df, xlsx_path)
    action_log.to_csv(csv_path, index=False, encoding="utf-8-sig")

    # 控制台摘要
    print("[DONE]")
    print(f"timeline_png: {png_path}")
    print(f"action_log_xlsx: {xlsx_path}")
    print(f"action_log_csv: {csv_path}")
    print(f"resolved_rows: {len(action_log)}")
    print(f"batch_count: {0 if batch_summary.empty else batch_summary.shape[0]}")
    print(f"qa_rows: {0 if qa_df.empty else qa_df.shape[0]}")


if __name__ == "__main__":
    main()
