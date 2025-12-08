# -*- coding: utf-8 -*-
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# ======= 配置区 =======
if len(sys.argv) < 2:
    raise SystemExit("❌ 请在运行时指定数据源文件夹路径，例如：python qpcr_ingest.py /path/to/folder")

FOLDER = Path(sys.argv[1]).expanduser().resolve()
PATTERN = "*.txt"

MAP_FILE = FOLDER / "实验设计.xlsx"
TARGET_FILE = Path(r"04_data/raw/qpcr/qpcr_original_data.xlsx")

TXT_ENCODING = "utf-8-sig"

BLOCKS_PER_ROW = 10    # 每行放 10 个小矩阵
BLOCK_HEIGHT = 13      # 每个小矩阵固定行数
BLOCK_WIDTH = 3        # 每个小矩阵固定列数
# ===========================================

def read_txt(path: Path) -> pd.DataFrame:
    """读取 txt 数据，返回 Samples / MeanCp / STD Cp 三列"""
    df = pd.read_csv(
        path,
        sep="\t",
        engine="python",
        encoding=TXT_ENCODING,
        skiprows=1,
        header=0
    )
    df = df[['Samples', 'MeanCp', 'STD Cp']]

    # 补齐到至少 BLOCK_HEIGHT - 1 行（BLOCK_HEIGHT - 1 表头行 + 数据行）
    # 注意：如果原始行数多于 BLOCK_HEIGHT - 1，则不再截断，全部保留。
    if len(df) < (BLOCK_HEIGHT - 1):
        empty_rows = pd.DataFrame(
            [[""] * BLOCK_WIDTH] * ((BLOCK_HEIGHT - 1) - len(df)),
            columns=df.columns
        )
        df = pd.concat([df, empty_rows], ignore_index=True)

    return df

def load_mapping():
    """读取实验设计.xlsx 里的 gene 页作为映射表"""
    mapping_df = pd.read_excel(MAP_FILE, sheet_name="gene", dtype=str)

    # 清洗列名，避免空格/大小写问题
    mapping_df.columns = mapping_df.columns.str.strip().str.lower()

    mapping_df['serial_number'] = mapping_df['serial_number'].str.strip().str.lower()
    mapping_df['gene_name'] = mapping_df['gene_name'].str.strip()
    return dict(zip(mapping_df['serial_number'], mapping_df['gene_name']))

def find_first_empty_row_in_col(ws, col=1):
    """扫描指定列（默认 A 列）找第一个空行"""
    for r in range(1, ws.max_row + 2):
        if ws.cell(row=r, column=col).value in (None, ""):
            return r
    return ws.max_row + 1

def is_full_row_repeated_header(row_values, block_width=3):
    """判断整行是否为重复的表头（横向拼接的 Samples/MeanCp/STD Cp）"""
    header_block = ["Samples", "MeanCp", "STD Cp"]
    return (
        len(row_values) % block_width == 0
        and all(list(row_values[i:i+block_width]) == header_block
                for i in range(0, len(row_values), block_width))
    )

def main():
    mapping = load_mapping()

    files = sorted(FOLDER.glob(PATTERN), key=lambda p: int(''.join(filter(str.isdigit, p.stem)) or 0))
    if not files:
        raise SystemExit(f"未找到 txt 文件：{FOLDER}")

    wb = load_workbook(TARGET_FILE)
    ws = wb.worksheets[0]

    # 起始行、列
    start_row = find_first_empty_row_in_col(ws, col=1)
    start_col = 8  # H 列

    for idx, p in enumerate(files):
        serial_number = ''.join(filter(str.isdigit, p.stem)).lower()
        gene_name = mapping.get(serial_number, p.stem)

        df = read_txt(p)

        # 计算块位置
        row_block = idx // BLOCKS_PER_ROW
        col_block = idx % BLOCKS_PER_ROW
        block_start_row = start_row + row_block * BLOCK_HEIGHT
        block_start_col = start_col + col_block * BLOCK_WIDTH

        # 第一行写 gene_name
        for c in range(BLOCK_WIDTH):
            ws.cell(row=block_start_row, column=block_start_col + c, value=gene_name)


        # 数据行（去掉横向重复的表头）
        for r_idx, row in enumerate(df.itertuples(index=False), start=1):
            row_values = list(row)
            if is_full_row_repeated_header(row_values, BLOCK_WIDTH):
                continue
            for c_idx, value in enumerate(row_values, start=1):
                ws.cell(row=block_start_row + r_idx, column=block_start_col + c_idx - 1, value=value)

    wb.save(TARGET_FILE)
    print(f"✅ 数据已追加到 {TARGET_FILE}")

if __name__ == "__main__":
    main()